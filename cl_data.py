#author:ling
#date:2021.11.15

'''数据清洗'''

from openpyxl import load_workbook
import numpy as np

wb = load_workbook('lingyin.xlsx')
ws = wb['info']

rows = ws.max_row
cols = ws.max_column

ws.cell(1,9).value = "最低薪资"
ws.cell(1,10).value = "最高薪资"


for row in range(2,rows+1):
    s = ws.cell(row,2).value

    if s == '暂定':
        min_s = 0
        max_s = 0
    else:
        min_s = int(s.split("-")[0])
        max_s = int(s.split("-")[1])

    ws.cell(row,9).value = min_s/1000
    ws.cell(row,10).value = max_s/1000

wb.save('lingyin.xlsx')
print("finish!")

