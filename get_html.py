# author:ling
# date: 2021.11.12

from selenium import webdriver
import time
import re
from openpyxl import load_workbook
from scrapy import Selector

'''
步骤：
1、加载更多内容
2、获取所有内容
3、直接xpath找到链接内容
4、写入xlsx文件中

'''


def getInfo(html):

    '''xpath和re获取信息'''

    # 获取链接
    # 职业信息
    links_1 = re.findall(r'<a class="base-card__full-link" href=(.*?)data-tracking-control-name', html)
    links_2 = re.findall(r'data-tracking-control-name="public_jobs_jserp-result_search-card" href=(.*?)class="base-card__full-link">',html)
    links = links_1 + links_2
    time.sleep(3)
    print("finish")

    return links

def saveData(links):

    '''存储数据'''

    filename = './lingyin.xlsx'

    wb = load_workbook(filename)
    wb.create_sheet('link',0)
    ws = wb['link']


    ws.append([ '链接'])

    for i in range(len(links)):
        line = [links[i]]
        ws.append(line)
        # time.sleep(self.timeout)

    wb.save(filename)


if __name__ == '__main__':
    # driver设置
    driver = webdriver.Chrome()
    # 浏览器页面
    driver.get('https://www.linkedin.com/jobs/search?keywords=%E7%AE%97%E6%B3%95%E5%B7%A5%E7%A8%8B%E5%B8%88&location=%E5%85%A8%E7%90%83%E8%8C%83%E5%9B%B4&geoId=92000000&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0')

    Links = []
    for i in range(150):

        driver.execute_script("window.scrollTo(0, document.body.clientHeight);")  # 下滑一个窗口高度
        time.sleep(1)

    # 获取所有页面信息
    # 异步加载的数据需要等待时长获取信息
    waittime = 4
    time.sleep(waittime)

    # 获取信息
    html = driver.page_source

    # xpath索取
    links = getInfo(html)

    for n in range(len(links)):
        Links.append(links[n].replace('"',''))

    # 存储信息
    saveData(Links)

    driver.quit()