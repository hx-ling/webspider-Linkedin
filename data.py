# author:ling
# date:2021.11.13

'''处理数据'''

from openpyxl import load_workbook
import numpy as np

def subStr(s):
    s = s.replace("[", "")
    s = s.replace("]", "")
    s = s.strip("'")
    return s

def subStr2(s):
    s = s.replace(" ", "")
    s = s.replace("xa0","")
    s = s.replace(",","")
    s = s.replace("[", "")
    s = s.replace("]", "")
    s = s.strip("'")
    return s

def subStr3(s):
    s = s.replace(" ", "")
    s = s.replace("</p>","")
    s = s.replace("<p>", "")

    s = s.replace("<br>","")
    s = s.replace("<ul>","")
    s = s.replace("</ul>","")
    s = s.replace("<li>","")
    s = s.replace("</li>", "")

    s = s.replace("</div>","")
    s = s.replace("<strong>", "")
    s = s.replace("</strong>", "")

    s = s.replace("<em>","")
    s = s.replace("</em>","")

    s = s.replace("<u>","")
    s = s.replace("</u>","")

    s = s.replace(",","")
    s = s.replace("[", "")
    s = s.replace("]", "")
    s = s.strip("'")

    return s

def st_split(s):
    st = s.split("要求")
    return st

# open workbook
wb = load_workbook('lingyin.xlsx')
ws = wb['info']

rows = ws.max_row
cols = ws.max_column

for row in range(1,rows+1):
    ws.cell(row,1).value = subStr(ws.cell(row,1).value)
    ws.cell(row,2).value = subStr2(ws.cell(row,2).value)
    ws.cell(row,3).value = subStr(ws.cell(row,3).value.replace(", 中国",""))
    ws.cell(row,4).value = subStr(ws.cell(row,4).value)
    ws.cell(row,5).value = subStr(ws.cell(row,5).value)
    ws.cell(row,6).value = subStr3(ws.cell(row,6).value)
    ws.cell(row,7).value = st_split(ws.cell(row,6).value)[0]
    ws.cell(row,8).value = st_split(ws.cell(row,6).value)[-1]

# 保存
wb.save('./lingyin.xlsx')
print("save!")