# author:ling
# date:2021.11.12

import re
import pandas as pd
import urllib.request
import chardet
import random
from openpyxl import load_workbook
import time
from selenium import webdriver
from scrapy import Selector

class Lingyin(object):

    def __init__(self):
        # 读入xlsx文件中的链接
        data = pd.read_excel('lingyin.xlsx')
        self.urls = data.iloc[:, 0]

        self.job_name = []  # 职业名称
        self.salary = []  # 薪资
        self.area = []  # 地区

        self.com_name = []  # 公司名称
        self.com_tag = []  # 公司类别

        self.job_need = []  # 岗位职责

        self.timeout = 2
        self.filename = 'lingyin.xlsx'

        self.getInfo()
        self.saveInfo()

    def subStr(self,st):
        str = []
        if st == []:
            str.append("暂定")
        else:
            for s in st:
                s = s.replace("\n", "")
                s = s.replace("[", "")
                s = s.replace("]","")
                s = s.replace('<div class="show-more-less-html__markup show-more-less-html__markup--clamp-after-5">','')
                s = s.strip("'")
                str.append(s)
        return str

    def getInfo(self):
        for i in range(len(self.urls)):
            # 获取页面html

            # driver设置
            driver = webdriver.Chrome()
            # 浏览器页面
            driver.get(self.urls[i])

            # 异步加载的数据需要等待时长获取信息
            waittime = 1
            loadmore = False
            time.sleep(waittime)

            # 获取信息
            html = driver.page_source
            selector = Selector(text=html)

            job_name = selector.xpath('//h3[@class="sub-nav-cta__header"]/text()').extract()
            job_name = self.subStr(job_name)

            salary = selector.xpath('//div[@class="salary compensation__salary"]/text()').extract()
            salary = self.subStr(salary)

            area = selector.xpath('//span[@class="sub-nav-cta__meta-text"]/text()').extract()
            area = self.subStr(area)

            com_name = re.findall(r'class="sub-nav-cta__optional-url" title="(.*?)" data-tracking-control-name', html)
            com_name = self.subStr(com_name)

            com_tag = re.findall(r'<a href="(.*?)" class="sub-nav-cta__optional-url"',html)
            com_tag = self.subStr(com_tag)

            # job和need的re索引有问题，因此该部分换成scrapy爬虫

            job_need = selector.xpath('//div[@class="show-more-less-html__markup show-more-less-html__markup--clamp-after-5"]').extract()
            job_need = self.subStr(job_need)

            time.sleep(self.timeout)
            print("%d finish" % (i+1))

            # 填入数组之中
            self.job_name.append(job_name)
            self.salary.append(salary)
            self.area.append(area)

            self.com_name.append(com_name)
            self.com_tag.append(com_tag)

            self.job_need.append(job_need)

            driver.quit()

    def saveInfo(self):
        wb = load_workbook(self.filename)
        wb.create_sheet('info',1)
        ws = wb['info']

        ws.append(['职业名称','薪水', '地区', '公司名称', '公司链接', '岗位职责和需求'])

        for i in range(len(self.salary)):
            line = [str(self.job_name[i]),str(self.salary[i]), str(self.area[i]), str(self.com_name[i]),
                    str(self.com_tag[i]),str(self.job_need[i])]
            ws.append(line)
            # time.sleep(self.timeout)

        wb.save(self.filename)
        print("save!")


if __name__ == '__main__':
    lingyin = Lingyin()